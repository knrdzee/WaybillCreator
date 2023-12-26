import sys
import json
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QComboBox, QVBoxLayout, QPushButton, QFormLayout, QDialog, \
    QTableWidget, QTableWidgetItem, QDoubleSpinBox, QTextEdit, QMessageBox, QLineEdit, QFileDialog, QInputDialog
import traceback
from datetime import datetime
import pandas as pd
import os
from PIL import Image
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.reader.excel import load_workbook


class LoginDialog(QDialog):
    def __init__(self):
        super(LoginDialog, self).__init__()

        self.setWindowTitle('Вход')
        self.setGeometry(300, 300, 300, 150)

        self.init_ui()

    def init_ui(self):
        self.label_username = QLabel('Логин:')
        self.edit_username = QLineEdit(self)

        self.label_password = QLabel('Пароль:')
        self.edit_password = QLineEdit(self)
        self.edit_password.setEchoMode(QLineEdit.Password)

        self.button_login = QPushButton('Войти', self)
        self.button_login.clicked.connect(self.login)

        layout = QVBoxLayout(self)
        layout.addWidget(self.label_username)
        layout.addWidget(self.edit_username)
        layout.addWidget(self.label_password)
        layout.addWidget(self.edit_password)
        layout.addWidget(self.button_login)

    def login(self):
        # Здесь вы можете реализовать проверку логина и пароля
        username = self.edit_username.text()
        password = self.edit_password.text()

        # проверка логина и пароля
        if username == 'polzovatel' or username == 'Polzovatel' and password == '1':
            self.accept()  # Подтверждаем вход
            QMessageBox.information(self, 'Успешная авторизация', 'Авторизация прошла успешно', QMessageBox.Ok)
        else:
            QMessageBox.warning(self, 'Ошибка', 'Логин и пароль отклонены, авторизация не пройдена', QMessageBox.Ok)

class EditInvoiceApp(QDialog):
    def __init__(self, loaded_invoice_data, json_file_path):
        super(EditInvoiceApp, self).__init__()

        self.setWindowTitle('Редактирование накладной')
        self.setGeometry(100, 100, 600, 400)

        self.loaded_invoice_data = loaded_invoice_data
        self.json_file_path = json_file_path  # Добавляем путь к файлу JSON

        self.init_ui()

    def init_ui(self):
        vbox_layout = QVBoxLayout()

        self.data_label = QLabel('Данные из накладной:')
        self.data_edit = QTextEdit()
        self.data_edit.setPlainText(json.dumps(self.loaded_invoice_data, ensure_ascii=False, indent=4))
        self.data_edit.setReadOnly(False)

        vbox_layout.addWidget(self.data_label)
        vbox_layout.addWidget(self.data_edit)

        self.edit_button = QPushButton('Сохранить изменения', self)
        self.edit_button.clicked.connect(self.save_changes)

        self.export_button = QPushButton('Экспортировать в Excel', self)
        self.export_button.clicked.connect(self.convert_to_excel)


        vbox_layout.addWidget(self.edit_button)
        vbox_layout.addWidget(self.export_button)

        self.setLayout(vbox_layout)

    def save_changes(self):
        try:
            edited_data = json.loads(self.data_edit.toPlainText())
            # Сохраняем измененные данные в оригинальной переменной

            self.loaded_invoice_data = edited_data

            # Сохраняем изменения в исходный файл JSON
            with open(self.json_file_path, 'w', encoding='utf-8') as json_file:
                json.dump(edited_data, json_file, ensure_ascii=False, indent=4)

            # Просто выведем сообщение, что изменения сохранены
            QMessageBox.information(self, 'Изменения сохранены', 'Изменения в накладной сохранены', QMessageBox.Ok)
        except Exception as e:
            print(f"Ошибка при сохранении изменений: {e}")
            traceback.print_exc()
            QMessageBox.warning(self, 'Ошибка', f'Не удалось сохранить изменения: {str(e)}', QMessageBox.Ok)

    def add_stamp_to_excel(self, excel_file_path, stamp_image_path):
        # Открываем изображение с печатью
        stamp_img = Image.open(stamp_image_path)

        # Сохраняем изображение во временный файл
        temp_stamp_path = "temp_stamp.png"
        stamp_img.save(temp_stamp_path)

        try:
            # Открываем существующий файл Excel
            workbook = load_workbook(excel_file_path)
            sheet = workbook.active

            # Создаем объект ExcelImage и добавляем его в лист
            excel_image = ExcelImage(temp_stamp_path)
            sheet.add_image(excel_image, 'I1')

            # Сохраняем изменения
            workbook.save(excel_file_path)
        except Exception as e:
            print(f"Ошибка при добавлении печати: {e}")
        finally:
            # Удаляем временный файл
            os.remove(temp_stamp_path)

    def convert_to_excel(self):
        try:
            # Развертываем структуру JSON в DataFrame
            df_items = pd.json_normalize(self.loaded_invoice_data, 'Товары', sep='_')

            # Создаем DataFrame для остальных данных (Магазин, Склад, Поставщик, Общая стоимость закупки)
            df_metadata = pd.DataFrame([self.loaded_invoice_data])

            # Объединяем два DataFrame
            df = pd.concat([df_metadata, df_items], axis=1)

            # Сохраняем значение 'Общая стоимость закупки'
            total_purchase_cost = df['Общая стоимость закупки'][0]

            # Убираем дублирующиеся столбцы
            df.drop(columns=['Общая стоимость закупки'], inplace=True)

            # Переупорядочиваем столбцы
            column_order = ['Магазин', 'Склад', 'Поставщик', 'Наименование', 'Измерение', 'Количество', 'Общая цена']
            df = df[column_order]

            # Открываем диалог сохранения
            excel_file, _ = QFileDialog.getSaveFileName(self, 'Сохранить как Excel', filter='Excel Files (*.xlsx)')

            if excel_file:
                # Сохраняем в Excel
                df.to_excel(excel_file, index=False)

                # Добавляем 'Общая стоимость закупки' в конце
                with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']
                    row, col = df.shape
                    worksheet.cell(row=row + 2, column=col + 1, value='Общая стоимость закупки')
                    worksheet.cell(row=row + 3, column=col + 1, value=total_purchase_cost)

                self.add_stamp_to_excel(excel_file, "stamp/SHablon-IP-02.png")

                QMessageBox.information(self, 'Экспорт завершен', 'Данные успешно экспортированы в Excel',
                                        QMessageBox.Ok)

        except Exception as e:
            print(f"Ошибка при экспорте в Excel: {e}")
            traceback.print_exc()
            QMessageBox.warning(self, 'Ошибка', f'Не удалось экспортировать данные: {str(e)}', QMessageBox.Ok)


# Добавим кнопку для загрузки накладной в основное окно (InvoiceApp)

class InvoiceApp(QDialog):
    def __init__(self):
        super(InvoiceApp, self).__init__()

        self.setWindowTitle('Создание накладной')
        self.setGeometry(100, 100, 800, 600)

        # Ваши данные о товарах, их ценах, измерениях и адресах магазинов
        self.product_prices = {
            'Молоко Простоквашино': {'price': 50, 'measurement': 'л'},
            'Хлеб Бородинский': {'price': 25, 'measurement': 'шт'},
            'Яйца Деревенские': {'price': 90, 'measurement': 'шт'},
            'Картошка Роза': {'price': 30, 'measurement': 'кг'},
            'Масло Сливочное Веселая Корова': {'price': 100, 'measurement': 'шт'},
            'Сыр Российский': {'price': 60, 'measurement': 'кг'},
            'Йогурт Actimel': {'price': 55, 'measurement': 'л'},
            'Курица Деревенская': {'price': 80, 'measurement': 'кг'},
            'Помидоры Агромаркет': {'price': 25, 'measurement': 'кг'},
            'Бекон Докторской': {'price': 90, 'measurement': 'шт'},
            'Крупа Геркулес': {'price': 40, 'measurement': 'кг'},
            'Сок Яблочный Фруктовый Сад': {'price': 50, 'measurement': 'л'},
            'Паста Barilla': {'price': 20, 'measurement': 'кг'},
            'Яблоки Голден': {'price': 30, 'measurement': 'кг'},
            'Чипсы Lay\'s': {'price': 65, 'measurement': 'шт'},
            'Лосось Норвежский': {'price': 85, 'measurement': 'кг'},
            'Кофе Арабика': {'price': 70, 'measurement': 'г'},
            'Мороженое Бриошь': {'price': 55, 'measurement': 'л'},
            'Пицца Ristorante': {'price': 75, 'measurement': 'шт'},
            'Мед Майский': {'price': 40, 'measurement': 'кг'},
            'Чай Ahmad Tea': {'price': 30, 'measurement': 'г'},
            'Апельсиновый сок Rich': {'price': 60, 'measurement': 'л'},
            'Рис Uncle Ben\'s': {'price': 25, 'measurement': 'кг'},
            'Огурцы Зеленые': {'price': 20, 'measurement': 'кг'},
            'Шоколад Аленка': {'price': 90, 'measurement': 'шт'},
            'Печенье Мария': {'price': 35, 'measurement': 'шт'},
            'Масло оливковое Bertolli': {'price': 80, 'measurement': 'л'},
            'Клубника Свежая': {'price': 45, 'measurement': 'кг'},
            'Вино Красное': {'price': 65, 'measurement': 'л'},
            'Лимонад Мираторг': {'price': 25, 'measurement': 'л'},
        }

        self.store_addresses = {
            'Золотой ключик': 'ул.Смирнова, д. 72/3',
            'Колокольчик': 'ул.Багаева, д. 12',
            'Магнит': 'ул.Куконковых, д. 132',
            'Пятёрочка': 'ул.Пушкина, д. 90',
            'Перекрёсток': 'ул.8 Марта, д. 32',
            'ВкусВилл': 'пр.Ленина, д. 39',
            'Лента': 'ул.Карла Маркса, д. 3',
            'Семейный': 'ул.Дзержинского, д. 27'
        }

        self.init_ui()

    def init_ui(self):
        self.setStyleSheet("background-color: white;")

        self.store_label = QLabel('Магазин:')
        self.store_edit = QComboBox()
        self.store_edit.addItems(self.store_addresses.keys())
        self.store_edit.currentIndexChanged.connect(self.update_address)

        self.address_label = QLabel('Адрес магазина:')
        self.address_edit = QLabel('')  # Поле для отображения адреса
        self.update_address()  # Установим адрес по умолчанию

        self.warehouse_label = QLabel('Склад:')
        self.warehouse_edit = QComboBox()
        self.warehouse_edit.addItems(['Центральный склад', 'Склад "ЮГ"', 'Склад "Важный"'])

        self.supplier_label = QLabel('Поставщик:')
        self.supplier_edit = QComboBox()
        suppliers = {'ИП Хачатрян Давид': '123', 'ООО "ЮГ"': '456', 'ИП Важный Игорь Витальевич': '789'}
        self.supplier_edit.addItems(suppliers.keys())

        self.products_label = QLabel('Товары:')
        self.products_table = QTableWidget()
        self.products_table.setColumnCount(5)
        self.products_table.setHorizontalHeaderLabels(['Товар', 'Цена за единицу', 'Измерение', 'Количество', 'Общая цена'])
        self.add_product_button = QPushButton('Добавить товар')
        self.add_product_button.clicked.connect(self.add_product)

        self.total_edit = QTextEdit()

        self.save_button = QPushButton('Сохранить накладную')
        self.save_button.clicked.connect(self.save_invoice)

        self.load_button = QPushButton('Загрузить накладную')
        self.load_button.clicked.connect(self.load_invoice)

        form_layout = QFormLayout()
        form_layout.addRow(self.store_label, self.store_edit)
        form_layout.addRow(self.address_label, self.address_edit)
        form_layout.addRow(self.warehouse_label, self.warehouse_edit)
        form_layout.addRow(self.supplier_label, self.supplier_edit)

        vbox_layout = QVBoxLayout()
        vbox_layout.addLayout(form_layout)
        vbox_layout.addWidget(self.products_label)
        vbox_layout.addWidget(self.products_table)
        vbox_layout.addWidget(self.add_product_button)
        vbox_layout.addWidget(self.total_edit)
        vbox_layout.addWidget(self.save_button)
        vbox_layout.addWidget(self.load_button)

        self.setLayout(vbox_layout)

    def update_address(self):
        selected_store = self.store_edit.currentText()
        self.address_edit.setText(self.store_addresses.get(selected_store, ''))

    def connect_quantity_changed(self, quantity_spinbox, row):
        quantity_spinbox.valueChanged.connect(lambda value, r=row: self.update_total(r))

    def add_product(self):
        row_position = self.products_table.rowCount()
        self.products_table.insertRow(row_position)

        product_item = QComboBox()
        product_item.addItems(self.product_prices.keys())
        price_item = QLabel('Цена')
        measurement_item = QLabel('Измерение')
        quantity_item = QDoubleSpinBox()
        quantity_item.setSingleStep(1.0)
        quantity_item.setValue(1.0)
        total_item = QLabel('Общая цена')

        product_item.currentIndexChanged.connect(lambda _, r=row_position: self.update_total(r))
        self.connect_quantity_changed(quantity_item, row_position)  # Подключаем обработчик к изменению количества

        self.products_table.setCellWidget(row_position, 0, product_item)
        self.products_table.setCellWidget(row_position, 1, price_item)
        self.products_table.setCellWidget(row_position, 2, measurement_item)
        self.products_table.setCellWidget(row_position, 3, quantity_item)
        self.products_table.setCellWidget(row_position, 4, total_item)

        self.update_total(row_position)

    def update_total(self, row):
        product_combo = self.products_table.cellWidget(row, 0)
        price_label = self.products_table.cellWidget(row, 1)
        measurement_label = self.products_table.cellWidget(row, 2)
        quantity_spinbox = self.products_table.cellWidget(row, 3)
        total_label = self.products_table.cellWidget(row, 4)

        self.connect_quantity_changed(quantity_spinbox, row)

        product_name = product_combo.currentText()
        product_info = self.product_prices.get(product_name, {'price': 0.0, 'measurement': 'ед'})
        price_label.setText(str(product_info['price']))
        measurement_label.setText(product_info['measurement'])
        quantity = quantity_spinbox.value()
        total = quantity * product_info['price']

        total_label.setText(str(total))
        self.calculate_total()

    def calculate_total(self):
        total = 0.0
        items_info = []

        for row in range(self.products_table.rowCount()):
            product_combo = self.products_table.cellWidget(row, 0)
            quantity_spinbox = self.products_table.cellWidget(row, 3)
            price_label = self.products_table.cellWidget(row, 1)
            total_label = self.products_table.cellWidget(row, 4)

            product_name = product_combo.currentText()
            product_info = self.product_prices.get(product_name, {'price': 0.0, 'measurement': 'ед'})
            quantity = quantity_spinbox.value()
            price = float(price_label.text())
            total = quantity * price

            total_label.setText(str(total))
            items_info.append(f"{product_name} - {quantity} {product_info['measurement']} - {total:.2f}")

        total = sum(
            float(self.products_table.cellWidget(row, 4).text()) for row in range(self.products_table.rowCount()))
        self.total_edit.setPlainText(f"Общая стоимость закупки: {total:.2f}\n" + '\n'.join(items_info))

    def save_invoice(self):
        try:
            data = {
                'Магазин': self.store_edit.currentText(),
                'Склад': self.warehouse_edit.currentText(),
                'Поставщик': self.supplier_edit.currentText(),
                'Товары': []
            }

            total_purchase_cost = 0.0

            for row in range(self.products_table.rowCount()):
                product_combo = self.products_table.cellWidget(row, 0)
                quantity_spinbox = self.products_table.cellWidget(row, 3)

                product_name = product_combo.currentText()
                product_info = self.product_prices.get(product_name, {'price': 0.0, 'measurement': 'ед'})
                quantity = quantity_spinbox.value()
                price = product_info['price']
                total_price = quantity * price

                total_purchase_cost += total_price

                data['Товары'].append({
                    'Наименование': product_name,
                    'Цена за единицу': price,
                    'Измерение': product_info['measurement'],
                    'Количество': quantity,
                    'Общая цена': total_price
                })

            data['Общая стоимость закупки'] = total_purchase_cost

            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")  # Уникальная временная метка
            filename = f'накладная_{timestamp}.json'

            with open(filename, 'w', encoding='utf-8') as json_file:
                json.dump(data, json_file, ensure_ascii=False, indent=4)

            QMessageBox.information(self, 'Сохранено', f'Накладная сохранена в файл "{filename}"', QMessageBox.Ok)
        except Exception as e:
            print(f"Ошибка при сохранении документа: {e}")
            traceback.print_exc()
            QMessageBox.warning(self, 'Ошибка', f'Не удалось сохранить накладную: {str(e)}', QMessageBox.Ok)

    

    def closeEvent(self, event):
        """Обработчик события закрытия окна."""
        reply = QMessageBox.question(self, 'Выход', 'Действительно ли вы хотите выйти?',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()



if __name__ == '__main__':
    app = QApplication(sys.argv)
    login_dialog = LoginDialog()
    if login_dialog.exec_() == QDialog.Accepted:
        invoice_app = InvoiceApp()
        invoice_app.show()
        sys.exit(app.exec_())
